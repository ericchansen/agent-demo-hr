"""export_roster: row_count reconciles with the scoped DB count; no sensitive cols."""

from __future__ import annotations

import csv

from openpyxl import load_workbook

from roster_mcp.db import connect
from roster_mcp.tools.count import count_roster
from roster_mcp.tools.export_roster import export_roster

SENSITIVE = {
    "gender",
    "age",
    "monthly_income",
    "salary_band",
    "performance_rating",
    "job_satisfaction",
    "environment_satisfaction",
}


def _scoped_active_count(scope) -> int:
    return count_roster({"active_status": "Active"}, scope)["count"]


def test_csv_export_rowcount_matches_scope(emea_scope):
    expected = _scoped_active_count(emea_scope)
    res = export_roster({"active_status": "Active"}, "csv", emea_scope)
    assert res["row_count"] == expected > 0
    with open(res["path"], newline="", encoding="utf-8") as fh:
        data_rows = list(csv.reader(fh))[1:]  # drop header
    assert len(data_rows) == expected


def test_xlsx_export_rowcount_matches_scope(emea_scope):
    expected = _scoped_active_count(emea_scope)
    res = export_roster({"active_status": "Active"}, "xlsx", emea_scope)
    assert res["row_count"] == expected
    wb = load_workbook(res["path"], read_only=True)
    ws = wb.active
    n_rows = sum(1 for _ in ws.iter_rows()) - 1  # minus header
    assert n_rows == expected


def test_export_excludes_sensitive_columns(emea_scope):
    res = export_roster({"active_status": "Active"}, "csv", emea_scope)
    assert SENSITIVE.isdisjoint(res["columns"])


def test_back_to_back_exports_use_distinct_paths(emea_scope):
    first = export_roster({"active_status": "Active"}, "csv", emea_scope)
    second = export_roster({"active_status": "Active"}, "csv", emea_scope)
    assert first["path"] != second["path"]


def test_export_never_leaks_out_of_scope(emea_scope):
    res = export_roster({}, "csv", emea_scope)
    with open(res["path"], newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    assert rows and all(r["region"] == "EMEA" for r in rows)
    # Cross-check against a known APAC employee id — must be absent.
    with connect() as conn:
        apac_id = conn.execute(
            "SELECT employee_id FROM fact_employee WHERE region='APAC' LIMIT 1"
        ).fetchone()[0]
    assert str(apac_id) not in {r["employee_id"] for r in rows}
